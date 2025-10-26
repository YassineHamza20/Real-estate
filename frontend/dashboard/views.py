# dashboard/views.py
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count, Q
from django.http import JsonResponse
from users.models import User, SellerVerification
from properties.models import Property, Wishlist
from datetime import datetime, timedelta
import json
from django.http import HttpResponse
import io
from io import BytesIO
from django.db.models import Avg, Max, Min, Sum, Count
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from users.models import User, SellerVerification
 
@staff_member_required
def admin_dashboard(request):
    try:
        # Basic counts
        total_users = User.objects.count()
        total_buyers = User.objects.filter(role='buyer').count()
        total_sellers = User.objects.filter(role='seller').count()
        total_admins = User.objects.filter(role='admin').count()
        total_properties = Property.objects.count()
        total_wishlists = Wishlist.objects.count()
        
        # Seller verification
        verified_sellers = SellerVerification.objects.filter(status='approved').count()
        pending_verifications = SellerVerification.objects.filter(status='pending').count()
        unsubmitted_sellers = max(0, total_sellers - (verified_sellers + pending_verifications))
        
        # Property types distribution
        property_types_data = Property.objects.values('property_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        property_type_labels = [item['property_type'].title() for item in property_types_data]
        property_type_counts = [item['count'] for item in property_types_data]
        
        # Cities data
        cities_data = Property.objects.values('city').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        city_labels = [item['city'] or 'Unknown' for item in cities_data]
        city_counts = [item['count'] for item in cities_data]
        
        # Rooms distribution
        rooms_data = Property.objects.values('number_of_rooms').annotate(
            count=Count('id')
        ).order_by('number_of_rooms')
        
        room_labels = []
        room_counts = []
        for item in rooms_data:
            room_labels.append(f"{item['number_of_rooms']} Room{'s' if item['number_of_rooms'] > 1 else ''}")
            room_counts.append(item['count'])
        
        # Recent properties with wishlist counts
        recent_properties = Property.objects.select_related('seller').annotate(
            wishlist_count=Count('wishlisted_by')
        ).order_by('-created_at')[:10]
        
        # Most wishlisted properties
        most_wishlisted = Property.objects.annotate(
            wishlist_count=Count('wishlisted_by')
        ).filter(wishlist_count__gt=0).order_by('-wishlist_count')[:10]

        context = {
            'current_time': timezone.now().strftime('%H:%M:%S'),
            'total_users': total_users,
            'total_buyers': total_buyers,
            'total_sellers': total_sellers,
            'total_admins': total_admins,
            'total_properties': total_properties,
            'total_wishlists': total_wishlists,
            'verified_sellers': verified_sellers,
            'pending_verifications': pending_verifications,
            'unsubmitted_sellers': unsubmitted_sellers,
            
            # Chart data
            'property_types_labels': property_type_labels,
            'property_types_data': property_type_counts,
            'cities_labels': city_labels,
            'cities_data': city_counts,
            'rooms_labels': room_labels,
            'rooms_data': room_counts,
            
            # Table data
            'recent_properties': recent_properties,
            'most_wishlisted': most_wishlisted,
        }
        
        return render(request, 'dashboard/admin_dashboard.html', context)
    
    except Exception as e:
        print(f"Dashboard error: {e}")
        # Return basic context even if there's an error
        return render(request, 'dashboard/admin_dashboard.html', {
            'total_users': 0,
            'total_buyers': 0,
            'total_sellers': 0,
            'total_admins': 0,
            'total_properties': 0,
            'total_wishlists': 0,
            'verified_sellers': 0,
            'pending_verifications': 0,
            'unsubmitted_sellers': 0,
            'property_types_labels': [],
            'property_types_data': [],
            'cities_labels': [],
            'cities_data': [],
            'rooms_labels': [],
            'rooms_data': [],
            'recent_properties': [],
            'most_wishlisted': [],
        })

def dashboard_api(request):
    """Simple API endpoint for real-time data"""
    try:
        data = {
            'total_buyers': User.objects.filter(role='buyer').count(),
            'total_sellers': User.objects.filter(role='seller').count(),
            'total_admins': User.objects.filter(role='admin').count(),
            'total_properties': Property.objects.count(),
            'total_wishlists': Wishlist.objects.count(),
            'timestamp': timezone.now().isoformat()
        }
        return JsonResponse(data)
    except:
        return JsonResponse({'error': 'API temporarily unavailable'}, status=500)
    



#PDF fct ------------
def export_dashboard_pdf(request):
    """Export complete dashboard data as PDF"""
    try:
        # Create PDF buffer with better margins
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Professional Header
        title = Paragraph("REAL ESTATE ANALYTICS REPORT", styles['Title'])
        elements.append(title)
        
        # Report metadata
        date_str = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal'])
        elements.append(date_str)
        elements.append(Spacer(1, 25))
        
        # Get comprehensive data
        total_users = User.objects.count()
        total_buyers = User.objects.filter(role='buyer').count()
        total_sellers = User.objects.filter(role='seller').count()
        total_admins = User.objects.filter(role='admin').count()
        total_properties = Property.objects.count()
        total_wishlists = Wishlist.objects.count()
        
        # Price statistics
        price_stats = Property.objects.aggregate(
            avg_price=Avg('price'),
            max_price=Max('price'),
            min_price=Min('price')
        )
        
        # Seller verification stats
        verified_sellers = SellerVerification.objects.filter(status='approved').count()
        pending_verifications = SellerVerification.objects.filter(status='pending').count()
        
        # Property type distribution
        property_types = Property.objects.values('property_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Recent activity
        new_users_week = User.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        new_properties_week = Property.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        # EXECUTIVE SUMMARY SECTION
        elements.append(Paragraph("EXECUTIVE SUMMARY", styles['Heading2']))
        
        summary_data = [
            ['PLATFORM OVERVIEW', 'STATISTICS'],
            ['Total Users', f"{total_users:,}"],
            ['Active Buyers', f"{total_buyers:,}"],
            ['Verified Sellers', f"{verified_sellers:,}"],
            ['Total Properties', f"{total_properties:,}"],
            ['Wishlist Engagements', f"{total_wishlists:,}"],
            ['New Users (7 days)', f"{new_users_week:,}"],
            ['New Properties (7 days)', f"{new_properties_week:,}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # PRICE ANALYTICS SECTION
        elements.append(Paragraph("PRICE ANALYTICS", styles['Heading2']))
        
        price_data = [
            ['PRICE METRIC', 'AMOUNT (euros)'],
            ['Average Property Price', f"{price_stats['avg_price'] or 0:,.2f}"],
            ['Most Expensive Property', f"{price_stats['max_price'] or 0:,.2f}"],
            ['Most Affordable Property', f"{price_stats['min_price'] or 0:,.2f}"],
        ]
        
        price_table = Table(price_data, colWidths=[180, 120])
        price_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1fae5')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        elements.append(price_table)
        elements.append(Spacer(1, 20))
        
        # PROPERTY TYPE BREAKDOWN SECTION
        elements.append(Paragraph("PROPERTY TYPE DISTRIBUTION", styles['Heading2']))
        
        if property_types:
            property_type_data = [['PROPERTY TYPE', 'COUNT', 'PERCENTAGE']]
            total_props = sum(item['count'] for item in property_types)
            
            for item in property_types:
                percentage = (item['count'] / total_props * 100) if total_props > 0 else 0
                property_type_data.append([
                    item['property_type'].title(),
                    str(item['count']),
                    f"{percentage:.1f}%"
                ])
            
            property_type_table = Table(property_type_data, colWidths=[120, 60, 80])
            property_type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#faf5ff')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd6fe')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(property_type_table)
        else:
            elements.append(Paragraph("No property type data available", styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # RECENT PROPERTIES SECTION
        elements.append(Paragraph("RECENT PROPERTY LISTINGS", styles['Heading2']))
        
        recent_properties = Property.objects.select_related('seller').annotate(
            wishlist_count=Count('wishlisted_by')
        ).order_by('-created_at')[:15]
        
        if recent_properties:
            property_data = [['PROPERTY', 'CITY', 'PRICE (euros)', 'TYPE', 'ROOMS', 'WISHLISTS']]
            
            for prop in recent_properties:
                property_data.append([
                    prop.name[:20] + '...' if len(prop.name) > 20 else prop.name,
                    prop.city[:12] if prop.city else 'N/A',
                    f"{prop.price:,.2f}",
                    prop.property_type[:8].title(),
                    str(prop.number_of_rooms),
                    str(prop.wishlist_count)
                ])
            
            property_table = Table(property_data, colWidths=[100, 60, 70, 50, 40, 50])
            property_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fffbeb')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fed7aa')),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fffbeb')]),
            ]))
            elements.append(property_table)
        else:
            elements.append(Paragraph("No recent properties available", styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # SELLER VERIFICATION STATUS
        elements.append(Paragraph("SELLER VERIFICATION STATUS", styles['Heading2']))
        
        verification_data = [
            ['STATUS', 'COUNT'],
            ['Verified Sellers', str(verified_sellers)],
            ['Pending Verification', str(pending_verifications)],
            ['Total Sellers', str(total_sellers)],
        ]
        
        verification_table = Table(verification_data, colWidths=[150, 80])
        verification_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fecaca')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        elements.append(verification_table)
        
        # FOOTER
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("--- End of Report ---", styles['Normal']))
        elements.append(Paragraph("Confidential Business Document - For Internal Use Only", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        # Prepare response with better filename
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"RealEstate_Dashboard_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}")
    

# dashboard/views.py - UPDATE IMPORTS AT THE TOP
from django.db.models import Count, Q, Avg, Max, Min, Sum  # Make sure all are imported



#excel fct ------------
def export_dashboard_excel(request):
    """Export complete dashboard data as Excel"""
    try:
        # Create a workbook and add worksheets
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get comprehensive data
        total_users = User.objects.count()
        total_buyers = User.objects.filter(role='buyer').count()
        total_sellers = User.objects.filter(role='seller').count()
        total_admins = User.objects.filter(role='admin').count()
        total_properties = Property.objects.count()
        total_wishlists = Wishlist.objects.count()
        
        # Price statistics
        price_stats = Property.objects.aggregate(
            avg_price=Avg('price'),
            max_price=Max('price'),
            min_price=Min('price'),
            total_value=Sum('price')
        )
        
        # Seller verification stats
        verified_sellers = SellerVerification.objects.filter(status='approved').count()
        pending_verifications = SellerVerification.objects.filter(status='pending').count()
        
        # Property type distribution
        property_types = Property.objects.values('property_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Recent activity
        new_users_week = User.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        new_properties_week = Property.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        # Recent properties
        recent_properties = Property.objects.select_related('seller').annotate(
            wishlist_count=Count('wishlisted_by')
        ).order_by('-created_at')[:50]
        
        # Most wishlisted properties
        most_wishlisted = Property.objects.annotate(
            wishlist_count=Count('wishlisted_by')
        ).filter(wishlist_count__gt=0).order_by('-wishlist_count')[:20]
        
        # User list
        recent_users = User.objects.select_related('seller_verification').order_by('-date_joined')[:50]
        
        # STYLES
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2c5a77", end_color="2c5a77", fill_type="solid")
        subheader_font = Font(bold=True, color="2c5a77", size=11)
        subheader_fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # WORKSHEET 1: EXECUTIVE SUMMARY
        ws1 = wb.create_sheet("Executive Summary")
        
        # Title
        ws1.merge_cells('A1:D1')
        ws1['A1'] = "REAL ESTATE ANALYTICS DASHBOARD - EXECUTIVE SUMMARY"
        ws1['A1'].font = Font(bold=True, size=14, color="2c5a77")
        ws1['A1'].alignment = center_align
        
        ws1.merge_cells('A2:D2')
        ws1['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws1['A2'].font = Font(size=10, italic=True)
        ws1['A2'].alignment = center_align
        
        # Platform Overview
        ws1['A4'] = "PLATFORM OVERVIEW"
        ws1['A4'].font = subheader_font
        ws1['A4'].fill = subheader_fill
        
        overview_data = [
            ['Metric', 'Count', 'Additional Info', 'Value'],
            ['Total Users', total_users, 'New Users (7 days)', new_users_week],
            ['Buyers', total_buyers, 'Buyer Percentage', f"{(total_buyers/total_users*100) if total_users > 0 else 0:.1f}%"],
            ['Sellers', total_sellers, 'Verified Sellers', verified_sellers],
            ['Admins', total_admins, 'Pending Verifications', pending_verifications],
            ['Total Properties', total_properties, 'New Properties (7 days)', new_properties_week],
            ['Wishlist Saves', total_wishlists, 'Engagement Rate', f"{(total_wishlists/total_users*100) if total_users > 0 else 0:.1f}%"],
        ]
        
        for row_idx, row_data in enumerate(overview_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 5:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
        
        # Price Analytics
        ws1['A12'] = "PRICE ANALYTICS"
        ws1['A12'].font = subheader_font
        ws1['A12'].fill = subheader_fill
        
        price_data = [
            ['Metric', 'Value (euros)'],
            ['Average Property Price', price_stats['avg_price'] or 0],
            ['Most Expensive Property', price_stats['max_price'] or 0],
            ['Most Affordable Property', price_stats['min_price'] or 0],
            ['Total Platform Value', price_stats['total_value'] or 0],
        ]
        
        for row_idx, row_data in enumerate(price_data, start=13):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 13:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
        
        # Format price cells as currency
        for row in range(14, 18):
            cell = ws1.cell(row=row, column=2)
            cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # WORKSHEET 2: PROPERTY ANALYTICS
        ws2 = wb.create_sheet("Property Analytics")
        
        # Title
        ws2.merge_cells('A1:G1')
        ws2['A1'] = "PROPERTY ANALYTICS"
        ws2['A1'].font = Font(bold=True, size=14, color="2c5a77")
        ws2['A1'].alignment = center_align
        
        # Property Type Distribution
        ws2['A3'] = "PROPERTY TYPE DISTRIBUTION"
        ws2['A3'].font = subheader_font
        ws2['A3'].fill = subheader_fill
        
        type_data = [['Property Type', 'Count', 'Percentage']]
        total_props = sum(item['count'] for item in property_types)
        
        for item in property_types:
            percentage = (item['count'] / total_props * 100) if total_props > 0 else 0
            type_data.append([
                item['property_type'].title(),
                item['count'],
                f"{percentage:.1f}%"
            ])
        
        for row_idx, row_data in enumerate(type_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 4:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
        
        # Recent Properties
        ws2['A10'] = "RECENT PROPERTIES (Last 50)"
        ws2['A10'].font = subheader_font
        ws2['A10'].fill = subheader_fill
        
        property_headers = ['Property Name', 'City', 'Price (euros)', 'Type', 'Rooms', 'Size', 'Wishlists', 'Seller', 'Created Date', 'Status']
        
        for col_idx, header in enumerate(property_headers, start=1):
            cell = ws2.cell(row=11, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
            cell.border = border
            cell.alignment = center_align
        
        for row_idx, prop in enumerate(recent_properties, start=12):
            ws2.cell(row=row_idx, column=1, value=prop.name).border = border
            ws2.cell(row=row_idx, column=2, value=prop.city).border = border
            ws2.cell(row=row_idx, column=3, value=float(prop.price)).border = border
            ws2.cell(row=row_idx, column=4, value=prop.property_type.title()).border = border
            ws2.cell(row=row_idx, column=5, value=prop.number_of_rooms).border = border
            ws2.cell(row=row_idx, column=6, value=float(prop.size) if prop.size else 0).border = border
            ws2.cell(row=row_idx, column=7, value=prop.wishlist_count).border = border
            ws2.cell(row=row_idx, column=8, value=prop.seller.username).border = border
            ws2.cell(row=row_idx, column=9, value=prop.created_at.strftime('%Y-%m-%d')).border = border
            ws2.cell(row=row_idx, column=10, value='Active' if prop.is_available else 'Inactive').border = border
        
        # Format price column as currency
        for row in range(12, 12 + len(recent_properties)):
            cell = ws2.cell(row=row, column=3)
            cell.number_format = '#,##0.00'
        
        # Most Wishlisted Properties
        start_row = len(recent_properties) + 15
        ws2.cell(row=start_row, column=1, value="MOST WISHLISTED PROPERTIES").font = subheader_font
        ws2.cell(row=start_row, column=1).fill = subheader_fill
        
        wishlist_headers = ['Property Name', 'City', 'Price (euros)', 'Wishlist Count', 'Popularity Score']
        
        for col_idx, header in enumerate(wishlist_headers, start=1):
            cell = ws2.cell(row=start_row+1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="ec4899", end_color="ec4899", fill_type="solid")
            cell.border = border
            cell.alignment = center_align
        
        for row_idx, prop in enumerate(most_wishlisted, start=start_row+2):
            popularity = (prop.wishlist_count / total_users * 100) if total_users > 0 else 0
            ws2.cell(row=row_idx, column=1, value=prop.name).border = border
            ws2.cell(row=row_idx, column=2, value=prop.city).border = border
            ws2.cell(row=row_idx, column=3, value=float(prop.price)).border = border
            ws2.cell(row=row_idx, column=4, value=prop.wishlist_count).border = border
            ws2.cell(row=row_idx, column=5, value=f"{popularity:.2f}%").border = border
        
        # Format price column as currency
        for row in range(start_row+2, start_row+2 + len(most_wishlisted)):
            cell = ws2.cell(row=row, column=3)
            cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths for ws2
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 30)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # WORKSHEET 3: USER ANALYTICS
        ws3 = wb.create_sheet("User Analytics")
        
        # Title
        ws3.merge_cells('A1:E1')
        ws3['A1'] = "USER ANALYTICS"
        ws3['A1'].font = Font(bold=True, size=14, color="2c5a77")
        ws3['A1'].alignment = center_align
        
        # User Distribution
        ws3['A3'] = "USER DISTRIBUTION"
        ws3['A3'].font = subheader_font
        ws3['A3'].fill = subheader_fill
        
        user_dist_data = [
            ['Role', 'Count', 'Percentage'],
            ['Buyers', total_buyers, f"{(total_buyers/total_users*100) if total_users > 0 else 0:.1f}%"],
            ['Sellers', total_sellers, f"{(total_sellers/total_users*100) if total_users > 0 else 0:.1f}%"],
            ['Admins', total_admins, f"{(total_admins/total_users*100) if total_users > 0 else 0:.1f}%"],
        ]
        
        for row_idx, row_data in enumerate(user_dist_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 4:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
        
        # Recent Users
        ws3['A8'] = "RECENT USERS (Last 50)"
        ws3['A8'].font = subheader_font
        ws3['A8'].fill = subheader_fill
        
        user_headers = ['Username', 'Email', 'Role', 'Phone', 'Join Date', 'Verification Status']
        
        for col_idx, header in enumerate(user_headers, start=1):
            cell = ws3.cell(row=9, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="84cc16", end_color="84cc16", fill_type="solid")
            cell.border = border
            cell.alignment = center_align
        
        for row_idx, user in enumerate(recent_users, start=10):
            verification_status = 'Not Seller'
            if user.role == 'seller':
                if hasattr(user, 'seller_verification'):
                    verification_status = user.seller_verification.status.title()
                else:
                    verification_status = 'Not Submitted'
            
            ws3.cell(row=row_idx, column=1, value=user.username).border = border
            ws3.cell(row=row_idx, column=2, value=user.email).border = border
            ws3.cell(row=row_idx, column=3, value=user.role.title()).border = border
            ws3.cell(row=row_idx, column=4, value=user.phone_number or 'N/A').border = border
            ws3.cell(row=row_idx, column=5, value=user.date_joined.strftime('%Y-%m-%d')).border = border
            ws3.cell(row=row_idx, column=6, value=verification_status).border = border
        
        # Auto-adjust column widths for ws3
        for column in ws3.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 25)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Prepare response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"RealEstate_Analytics_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel Export Error: {e}")
        print(f"Error details: {error_details}")
        return HttpResponse(f"Error generating Excel: {str(e)}")